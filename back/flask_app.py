import os
from flask import Flask, jsonify, request, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_admin import Admin
from flask_admin.contrib.sqla import ModelView
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from flask_admin.form import Select2Widget
from wtforms import SelectField
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///bd.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'mysecretkey'
app.config['UPLOAD_FOLDER'] = '/uploads/'

db = SQLAlchemy(app)

# Models
class TDataSession(db.Model):
    __tablename__ = 'tdatasession'
    id = db.Column(db.Integer, primary_key=True)
    IdDataSession = db.Column(db.Integer)
    Time = db.Column(db.Integer)
    HR = db.Column(db.Float)
    SpO2 = db.Column(db.Float)
    O2 = db.Column(db.Float)
    O2set = db.Column(db.Float)
    CO2 = db.Column(db.Float)
    F = db.Column(db.Float)
    V = db.Column(db.Float)
    Ps = db.Column(db.Float)
    Pd = db.Column(db.Float)

    def serialize(self):
        return {
            'id': self.id,
            'IdDataSession': self.IdDataSession,
            'Time': self.Time,
            'HR': self.HR,
            'SpO2': self.SpO2,
            'O2': self.O2,
            'O2set': self.O2set,
            'CO2': self.CO2,
            'F': self.F,
            'V': self.V,
            'Ps': self.Ps,
            'Pd': self.Pd,
        }



class TGroup(db.Model):
    __tablename__ = 'tgroup'
    NGroup = db.Column(db.Integer, primary_key=True)
    Name = db.Column(db.String(50))

    def serialize(self):
        return{
            'NGroup' : self.NGroup,
            'Name' : self.Name
        }

class TPatien(db.Model):
    __tablename__ = 'tpatien'
    Nmedcard = db.Column(db.Integer, primary_key=True)
    Surname = db.Column(db.String(50))
    FirstName = db.Column(db.String(50))
    Patronymic = db.Column(db.String(50))
    DataOfBirth = db.Column(db.Date)
    Sex = db.Column(db.String(2))
    Height = db.Column(db.Integer)
    Weight = db.Column(db.Integer)
    Address = db.Column(db.String(100))
    PhoneN = db.Column(db.String(50))
    Email = db.Column(db.String(50))
    Password = db.Column(db.String(128))
    Diagnosis = db.Column(db.String(50))
    NSessionDiagn = db.Column(db.Integer)
    NSessionTreat = db.Column(db.Integer)
    NGroup = db.Column(db.Integer, db.ForeignKey('tgroup.NGroup'))
    group = db.relationship('TGroup', backref='patients')

    def serialize(self):
        return {
            'Nmedcard' : self.Nmedcard,
            'Surname' : self.Surname,
            'FirstName' : self.FirstName,
            'Patronymic' : self.Patronymic,
            'DataOfBirth' : self.DataOfBirth,
            'Sex' : self.Sex,
            'Height' : self.Height,
            'Weight' : self.Weight,
            'Address' : self.Address,
            'PhoneN' : self.PhoneN,
            'Email' : self.Email,
            'Diagnosis' : self.Diagnosis,
            'NSessionDiagn' : self.NSessionDiagn,
            'NSessionTreat' : self.NSessionTreat,
            'NGroup' : self.NGroup
        }

class TSession(db.Model):
    __tablename__ = 'tsession'
    Nmedcard = db.Column(db.Integer, db.ForeignKey('tpatien.Nmedcard'))
    patient = db.relationship('TPatien', backref='sessions')
    Namesession = db.Column(db.String(45))
    DateOfSession = db.Column(db.DateTime)
    DurationOfSession = db.Column(db.Integer)
    DurationOfPause = db.Column(db.Integer)
    IdDataSession = db.Column(db.Integer, primary_key=True)
    Note = db.Column(db.String(45))
    SessionPatternNum = db.Column(db.Integer)

    def serialize(self):
        return {
            'Nmedcard': self.Nmedcard,
            'Namesession': self.Namesession,
            'DateOfSession': self.DateOfSession.isoformat() if self.DateOfSession else None,
            'DurationOfSession': self.DurationOfSession,
            'DurationOfPause': self.DurationOfPause,
            'IdDataSession': self.IdDataSession,
            'Note': self.Note,
            'SessionPatternNum': self.SessionPatternNum,
        }

class Doctor(db.Model):
    __tablename__ = 'doctor'
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(100))
    email = db.Column(db.String(100), unique=True)
    password = db.Column(db.String(128))

    def set_password(self, password):
        self.password = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password, password)

    def serialize(self):
        return {
            'id': self.id,
            'full_name': self.full_name,
            'email': self.email,
        }

#admin panel settings
class TPatienView(ModelView):
    column_list = ('Nmedcard', 'Surname', 'FirstName', 'Patronymic', 'DataOfBirth', 'Sex', 'Height', 'Weight', 'Address', 'PhoneN', 'Email', 'Diagnosis', 'NSessionDiagn', 'NSessionTreat', 'group', 'Password')
    column_labels = {
        'Nmedcard': '№ мед.картки',
        'Surname': 'Прізвище',
        'FirstName': "Ім'я",
        'Patronymic': 'По батькові',
        'DataOfBirth': 'Дата народження',
        'Sex': 'Стать',
        'Height': 'Зріст',
        'Weight': 'Вага',
        'Address': 'Адреса',
        'Password':'Password',
        'PhoneN': 'Телефон',
        'Email': 'Електронна адреса',
        'Diagnosis': 'Діагноз',
        'NSessionDiagn': 'К-сть сеансів діагностики',
        'NSessionTreat': 'К-сть сеансів лікування',
        'NGroup': "№ групи"
    }
    form_columns = ('Nmedcard', 'Surname', 'FirstName', 'Patronymic', 'DataOfBirth', 'Sex', 'Height', 'Weight', 'Address', 'PhoneN', 'Email', 'Diagnosis', 'NSessionDiagn', 'NSessionTreat', 'group', 'Password')

    form_ajax_refs = {
        'group': {
            'fields': ['NGroup'],
            'page_size': 10,
            'placeholder': 'Оберіть групу'
        },
    }
    def on_form_prefill(self, form, id):
        form.Nmedcard.data = TPatien.query.count() + 1


class TGroupView(ModelView):
    column_list = ('NGroup', 'Name')
    column_labels = {
        'NGroup': '№ групи',
        'Name': 'Назва групи'
    }
    form_columns = ('NGroup', 'Name')

    def on_form_prefill(self, form, id):
        form.NGroup.data = TGroup.query.count() + 1

class TDataSessionView(ModelView):
    column_list = ('id', 'IdDataSession', 'Time', 'HR', 'SpO2', 'O2', 'O2set', 'CO2', 'F', 'V', 'Ps', 'Pd')
    column_labels = {
        'id': 'ID',
        'IdDataSession': 'ID сеансу',
        'Time': 'Час',
        'HR': 'ЧСС',
        'SpO2': 'SpO2',
        'O2': 'O2',
        'O2set': 'O2set',
        'CO2': 'CO2',
        'F': 'F',
        'V': 'V',
        'Ps': 'Ps',
        'Pd': 'Pd'
    }
    form_columns = ('IdDataSession', 'Time', 'HR', 'SpO2', 'O2', 'O2set', 'CO2', 'F', 'V', 'Ps', 'Pd')

    def on_form_prefill(self, form, id):
        form.IdDataSession.data = TDataSession.query.count() + 1

class TSessionView(ModelView):
    column_list = ('Nmedcard', 'Namesession', 'DateOfSession', 'DurationOfSession', 'DurationOfPause', 'IdDataSession', 'Note', 'SessionPatternNum')
    column_labels = {
        'Nmedcard': '№ мед.картки',
        'Namesession': 'Тип сеансу',
        'DateOfSession': 'Дата сеансу',
        'DurationOfSession': 'Тривалість сеансу',
        'DurationOfPause': 'Тривалість паузи',
        'IdDataSession': 'ID сеансу',
        'Note': 'Примітка',
        'SessionPatternNum': '№ схеми сеансу',
    }
    form_columns = ('patient', 'Namesession', 'DateOfSession', 'DurationOfSession', 'DurationOfPause', 'IdDataSession', 'Note', 'SessionPatternNum')


    form_ajax_refs = {
        'patient': {
            'fields': ['Nmedcard'],
            'page_size': 10,
            'placeholder': 'Оберіть пацієнта'
        },
    }
    def on_form_prefill(self, form, id):
        form.IdDataSession.data = TSession.query.count() + 1

class DoctorView(ModelView):
    column_list = ('id', 'full_name', 'email')
    column_labels = {
        'id': 'ID',
        'full_name': 'ФИО',
        'email': 'Email',
    }
    form_columns = ('full_name', 'email', 'password')

    def on_model_change(self, form, model, is_created):
        if is_created:
            model.set_password(form.password.data)

# Create tables
with app.app_context():
    db.create_all()

# Flask-Admin
admin = Admin(app)
admin.add_view(TDataSessionView(TDataSession, db.session))
admin.add_view(TGroupView(TGroup, db.session))
admin.add_view(TPatienView(TPatien, db.session))
admin.add_view(TSessionView(TSession, db.session))
admin.add_view(DoctorView(Doctor, db.session))

def create_sessions_xlsx_by_group(group_id):
    group_name = TGroup.query.get(group_id).Name
    filename = f'sessions_group_{group_id}.xlsx'
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write headers
    worksheet['A1'] = '№ групи'
    worksheet['B1'] = 'ПІБ'
    worksheet['C1'] = 'Дата сеансу'
    worksheet['D1'] = 'Тип сеансу'
    headers = ['HR', 'SpO2', 'O2', 'O2set', 'CO2', 'F', 'V', 'Ps', 'Pd']
    for i, header in enumerate(headers, start=6):
        col_letter = get_column_letter(i)
        worksheet[f'{col_letter}1'] = header

    # Write data
    row_num = 2
    patients = TPatien.query.filter_by(NGroup=group_id).all()
    for patient in patients:
        sessions = TSession.query.filter_by(Nmedcard=patient.Nmedcard).all()
        for session in sessions:
            data_sessions = TDataSession.query.filter_by(IdDataSession=session.IdDataSession).all()
            for data_session in data_sessions:
                worksheet[f'A{row_num}'] = group_id
                worksheet[f'B{row_num}'] = f'{patient.Surname} {patient.FirstName} {patient.Patronymic}'
                worksheet[f'C{row_num}'] = session.DateOfSession.strftime('%d.%m.%y %H:%M:%S')
                worksheet[f'D{row_num}'] = session.Namesession
                for i, header in enumerate(headers, start=6):
                    col_letter = get_column_letter(i)
                    worksheet[f'{col_letter}{row_num}'] = getattr(data_session, header)
                row_num += 1

    # Save and return file
    workbook.save(filename)
    return filename

from openpyxl import Workbook
from flask import current_app

def create_sessions_xlsx_by_medcard(nmedcard):
    patient = TPatien.query.get(nmedcard)
    if not patient:
        return None

    # Create workbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Set header row
    header_row = ['Дата сеансу', 'Тип сеансу', 'HR', 'SpO2', 'O2', 'O2set', 'CO2', 'F', 'V', 'Ps', 'Pd']
    worksheet.append(header_row)

    # Get sessions for patient
    sessions = TSession.query.filter_by(Nmedcard=nmedcard).all()
    for session in sessions:
        data_sessions = TDataSession.query.filter_by(IdDataSession=session.IdDataSession).all()
        for data_session in data_sessions:
            row_data = [session.DateOfSession.strftime('%d.%m.%y %H:%M:%S'), session.Namesession,
                        data_session.HR, data_session.SpO2, data_session.O2, data_session.O2set,
                        data_session.CO2, data_session.F, data_session.V, data_session.Ps, data_session.Pd]
            worksheet.append(row_data)

    # Set patient info at the top of the worksheet
    worksheet.insert_rows(1)
    worksheet['A1'] = f'ПІБ: {patient.Surname} {patient.FirstName} {patient.Patronymic}'
    worksheet.insert_rows(2)
    worksheet['A2'] = f'№ групи: {patient.NGroup}'
    worksheet.insert_rows(3)
    worksheet['A3'] = f'# Мед. карти: {patient.Nmedcard}'

    # Set column widths
    for column in worksheet.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        worksheet.column_dimensions[column_name].width = adjusted_width

    # Save and return file
    filename = f"{patient.Surname}_{patient.FirstName}_{patient.Patronymic}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    file_path = f'sessions_medcard_{nmedcard}.xlsx'
    workbook.save(file_path)
    return file_path


@app.route('/api/patients', methods=['GET'])
def get_all_patients():
    patients = TPatien.query.all()
    return jsonify([patient.serialize() for patient in patients])

# Редактирование пациента
@app.route('/api/patients/<int:nmedcard>', methods=['PUT'])
def update_patient(nmedcard):
    patient = TPatien.query.get(nmedcard)
    if patient:
        data = request.json
        if 'DataOfBirth' in data:
            data['DataOfBirth'] = datetime.fromtimestamp(data['DataOfBirth'])
        for key, value in data.items():
            setattr(patient, key, value)
        db.session.commit()
        return jsonify(patient.serialize())
    else:
        return jsonify({"error": "Patient not found"}), 404


# Удаление пациента
@app.route('/api/patients/<int:nmedcard>', methods=['DELETE'])
def delete_patient(nmedcard):
    patient = TPatien.query.get(nmedcard)
    if patient:
        db.session.delete(patient)
        db.session.commit()
        return jsonify({"message": "Patient deleted"})
    else:
        return jsonify({"error": "Patient not found"}), 404

# Добавление пациента
@app.route('/api/patients', methods=['POST'])
def add_patient():
    data = request.json
    if 'DataOfBirth' in data:
        data['DataOfBirth'] = datetime.fromtimestamp(data['DataOfBirth'])
    new_patient = TPatien(**data)
    db.session.add(new_patient)
    db.session.commit()
    return jsonify(new_patient.serialize()), 201

# Получение списка всех групп
@app.route('/api/groups', methods=['GET'])
def get_all_groups():
    groups = TGroup.query.all()
    return jsonify([group.serialize() for group in groups])

# Удаление группы
@app.route('/api/groups/<int:ngroup>', methods=['DELETE'])
def delete_group(ngroup):
    group = TGroup.query.get(ngroup)
    if group:
        db.session.delete(group)
        db.session.commit()
        return jsonify({"message": "Group deleted"})
    else:
        return jsonify({"error": "Group not found"}), 404

# Изменение названия группы
@app.route('/api/groups/<int:ngroup>', methods=['PUT'])
def update_group(ngroup):
    group = TGroup.query.get(ngroup)
    if group:
        data = request.json
        group.Name = data['Name']
        db.session.commit()
        return jsonify(group.serialize())
    else:
        return jsonify({"error": "Group not found"}), 404

# Добавление группы
@app.route('/api/groups', methods=['POST'])
def add_group():
    data = request.json
    new_group = TGroup(**data)
    db.session.add(new_group)
    db.session.commit()
    return jsonify(new_group.serialize()), 201

# Перенос пациента в другую группу
@app.route('/api/patients/<int:nmedcard>/transfer/<int:ngroup>', methods=['PUT'])
def transfer_patient(nmedcard, ngroup):
    patient = TPatien.query.get(nmedcard)
    group = TGroup.query.get(ngroup)
    if patient and group:
        patient.NGroup = ngroup
        db.session.commit()
        return jsonify(patient.serialize())
    else:
        return jsonify({"error": "Patient or group not found"}), 404

# Получение списка процедур по ID пользователя
@app.route('/api/patients/<int:nmedcard>/procedures', methods=['GET'])
def get_patient_procedures(nmedcard):
    patient = TPatien.query.get(nmedcard)
    if patient:
        sessions = TSession.query.filter_by(Nmedcard=nmedcard).all()
        results = []
        for session in sessions:
            data_sessions = TDataSession.query.filter_by(IdDataSession=session.IdDataSession).all()
            session_data = {
                'Nmedcard': session.Nmedcard,
                'Namesession': session.Namesession,
                'DateOfSession': session.DateOfSession,
                'DurationOfSession': session.DurationOfSession,
                'DurationOfPause': session.DurationOfPause,
                'IdDataSession': session.IdDataSession,
                'Note': session.Note,
                'SessionPatternNum': session.SessionPatternNum,
                'DataSessions': [data_session.serialize() for data_session in data_sessions]
            }
            results.append(session_data)
        return jsonify(results)
    else:
        return jsonify({"error": "Patient not found"}), 404

@app.route('/api/groups/<int:group_id>/sessions', methods=['GET'])
def get_group_sessions(group_id):
    filename = create_sessions_xlsx_by_group(group_id)
    return send_file(filename, as_attachment=True)


@app.route('/api/patients/<int:nmedcard>/sessions', methods=['GET'])
def get_patient_sessions(nmedcard):
    filename = create_sessions_xlsx_by_medcard(nmedcard)
    return send_file(filename, as_attachment=True)

@app.route('/api/register', methods=['POST'])
def register():
    data = request.get_json()
    full_name = data.get('fullName')
    email = data.get('email')
    password = data.get('password')

    if not full_name or not email or not password:
        return jsonify({'error': 'Input must be with data'}), 400

    existing_doctor = Doctor.query.filter_by(email=email).first()
    if existing_doctor:
        return jsonify({'error': 'Doctor with this email already registered'}), 400

    new_doctor = Doctor(full_name=full_name, email=email, password=password)

    db.session.add(new_doctor)
    db.session.commit()

    return jsonify({'success': 'Success'}), 201

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    student = Doctor.query.filter_by(email=data['email'], password=data['password']).first()
    if student:
        return jsonify({'status': 'success', 'message': 'Success', 'student': student.serialize()})
    else:
        return jsonify({'status': 'error', 'message': 'Error'})

@app.route('/api/doctor/<int:doctor_id>', methods=['GET'])
def get_doctor(doctor_id):
    doctor = Doctor.query.get(doctor_id)
    if not doctor:
        return jsonify({'error': 'Doctor not found'}), 404

    return jsonify(doctor.serialize())

@app.route('/api/patients/register', methods=['POST'])
def register_patient():
    data = request.get_json()
    new_patient = TPatien(
        Nmedcard=data['Nmedcard'],
        Surname=data['Surname'],
        FirstName=data['FirstName'],
        Patronymic=data['Patronymic'],
        DataOfBirth=datetime.utcfromtimestamp(data['DataOfBirth']), # Convert the Unix timestamp to a datetime object
        Sex=data['Sex'],
        Height=data['Height'],
        Weight=data['Weight'],
        Address=data['Address'],
        PhoneN=data['PhoneN'],
        Email=data['Email'],
        Password=data['Password'],
        NGroup=1
    )
    db.session.add(new_patient)
    db.session.commit()

    return jsonify({'message': 'New patient registered'}), 201

@app.route('/api/patients/login', methods=['POST'])
def login_patient():
    data = request.json
    student = TPatien.query.filter_by(Email=data['email'], Password=data['password']).first()
    if student:
        return jsonify({'status': 'success', 'message': 'Success', 'student': student.serialize()})
    else:
        return jsonify({'status': 'error', 'message': 'Error'})

@app.route('/api/patients/<int:nmedcard>', methods=['GET'])
def get_patient_by_medcard(nmedcard):
    patient = TPatien.query.get(nmedcard)
    if patient:
        return jsonify(patient.serialize())
    else:
        return jsonify({"error": "Error"}), 404


if __name__ == '__main__':
    app.run()
